from openpyxl import load_workbook
from pdfrw import PdfWriter

workbook = load_workbook('TEST_0000.xlsx', data_only=True)
worksheet = workbook.active

pw = PdfWriter('test_23.pdf')
ws_range = worksheet.iter_rows('A1:G28')
for row in ws_range:
    s = ''
    for cell in row:
        if cell.value is None:
            s += ' ' * 11
        else:
            s += str(cell.value).rjust(10) + ' '
    pw.writeLine(s)
pw.savePage()
pw.close()