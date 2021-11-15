#import openpyxl
from openpyxl import Workbook
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder

wb = Workbook()
ws = wb.active

data = [
    ["Fruit", "Quantity"],
    ["Kiwi", 4],
    ["Grape", 5],
    ["Apple", 6],
    ["Peach", 7],
    ["Pomegranate", 10],
    ["Pear", 15],
    ["Tangerine",11],
    ["Blueberry", 11],
    ["Mango", 3],
    ["Watermelon", 5],
    ["Blackberry", 6],
    ["Orange", 7],
    ["Raspberry", 7],
    ["Banana", 3]
]

for r in data:
    ws.append(r)

ws.auto_filter.ref = "A1:B15"
#ws.auto_filter.add_filter_column(0, ["Kiwi", "Apple", "Mango"])
#ws.auto_filter.add_sort_condition("B2:B15")
print(ws.max_column)
print(ws.max_row)
wb.save("filtered.xlsx")