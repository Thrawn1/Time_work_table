from openpyxl import Workbook
from openpyxl.styles import Border, Side,Alignment,Font
from id_employee import id_employee
from analyze_data import month_name_for_print


def build_file_excel(time_table:dict,work_time_employees:dict,working_hours_of_workers_sum_of_all_data:dict,total_salary_id:dict):
    """Функция строит таблицу в excel файле. В функцию передаются: структура данных, содержащих данные об отметках и всех работниках за месяц,
    структуру данных, содержащих информацию о переработках в каждом дне месяца, для каждого работника, структуру данных, содержащую информацию
    за месяц для каждого работника, структуру, содержащую информацию про зарплату(если введен секретный ключ)
    Создает файл, внутрь внего записываются данные из структур и создаются границы. После этого файл сохраняется с именем, указывающим на месяц и год
    за кторый взяты данные. Имя созданного файла выводиться на экран, если файл создан.
    """

    
    wb = Workbook()
    list_employee = id_employee(type_data=2)
    ws = wb.active
    ws.column_dimensions['A'].width = 16.44 #Ширина столбцов в условных единицах табличного процессора(1 единица - 19,58 мм)
    ws.column_dimensions['B'].width = 13.48
    ws.column_dimensions['C'].width = 14.3
    ws.column_dimensions['D'].width = 16.43
    ws.column_dimensions['E'].width = 20.84
    ws.column_dimensions['F'].width = 12.26
    ws.column_dimensions['G'].width = 13.23
    ws.column_dimensions['H'].width = 16.44
    ws.column_dimensions['I'].width = 23.21
    ws.column_dimensions['J'].width = 12.46
    ws.column_dimensions['K'].width = 17.93
    ws.column_dimensions['L'].width = 21.96
    ws.column_dimensions['M'].width = 28.09
    ws.column_dimensions['N'].width = 10.27
    ws.column_dimensions['P'].width = 10.3  
    border_style='thin'
    color='FF000000'
    top = Side(border_style,color)
    right = Side(border_style,color)
    bottom = Side(border_style,color)
    left = Side(border_style,color)
    font_text_headline = Font(name='Calibri',size=11,bold=True)
    topicsList=['Фамилия', 'Дата', 'Отметка входа', 'Отметка выхода', 'Общее время работы', 'Переработка']
    topicCounter=1
    for topic in topicsList:
        ws.cell(column = topicCounter, row = 1, value = topic)
        ws.cell(1,topicCounter).font = font_text_headline
        ws.cell(1,topicCounter).border = Border(left,right,top,bottom)
        topicCounter+=1
    ws.merge_cells(start_row=1,start_column=6,end_row=1,end_column=7)
    ws.cell(column = 6,row = 1).alignment = Alignment(horizontal='center')
    count = 2
    list_date = []
    for date in time_table.keys():
        list_date.append(date)
    list_date.sort()
    for date in list_date:
        for id in time_table[date]:
            if id != 8 and id !=7:
                ws.cell(column = 1,row = count, value=list_employee[1][id])
                ws.cell(count,1).border = Border(left,right,top,bottom)
                ws.cell(column = 2, row = count, value = date)
                ws.cell(count,2).border = Border(left,right,top,bottom)
                if time_table[date][id][2] == 'work' or time_table[date][id][2] == 'weekend':
                    ws.cell(column = 3,row = count, value=time_table[date][id][1].time())
                    ws.cell(count,3).border = Border(left,right,top,bottom)
                    ws.cell(column = 4,row = count, value=time_table[date][id][0].time())
                    ws.cell(count,4).border = Border(left,right,top,bottom)
                    ws.cell(column = 5,row = count, value=work_time_employees[date][id][1])
                    ws.cell(count,5).border = Border(left,right,top,bottom)
                    ws.cell(column = 7,row = count, value=work_time_employees[date][id][0])
                    ws.cell(count,7).border = Border(left,right,top,bottom)
                    if work_time_employees[date][id][2] == 'недоработка':
                        ws.cell(column = 6,row = count, value=work_time_employees[date][id][2])
                        ws.cell(count,6).border = Border(left,right,top,bottom)
                    else:
                        ws.cell(column = 6,row = count, value=work_time_employees[date][id][2])
                        ws.cell(count,6).border = Border(left,right,top,bottom)
                elif time_table[date][id][2] == 'vacation':
                    ws.cell(count,3).border = Border(left,right,top,bottom) # Сначала границы, потом объединение. Иначе граница будет не верной
                    ws.merge_cells(start_row=count,start_column=3,end_row=count,end_column=7)
                    ws.cell(column = 3,row = count, value='Отпуск')
                    ws.cell(column = 3,row = count).alignment = Alignment(horizontal='center')
                elif time_table[date][id][2] == 'truancy':
                    ws.cell(count,3).border = Border(left,right,top,bottom)
                    ws.merge_cells(start_row=count,start_column=3,end_row=count,end_column=7)
                    ws.cell(column = 3,row = count, value='Прогул')
                    ws.cell(column = 3,row = count).alignment = Alignment(horizontal='center')
                else:
                    pass
            if id != 8 and id != 7:
                count += 1
    count += 3
    topicsList=['Фамилия', 'Отработано будних день', 'Переработка', 'Рабочих выходных','Переработка выходных', 'Количество дней дней отпуска', 'Оклад','Молоко','Зарплата']
    topicCounter=8
    for topic in topicsList:
        ws.cell(column = topicCounter, row = count, value = topic)
        ws.cell(count,topicCounter).border = Border(left,right,top,bottom)
        topicCounter+=1
    count += 1
    for id in working_hours_of_workers_sum_of_all_data.keys():
        ws.cell(column = 8,row = count, value=list_employee[1][id])
        ws.cell(count,8).border = Border(left,right,top,bottom)
        number_worked_day = working_hours_of_workers_sum_of_all_data[id][0][0]
        ws.cell(column = 9,row = count, value=number_worked_day)
        ws.cell(column = 9,row = count).alignment = Alignment(horizontal='center')
        ws.cell(count,9).border = Border(left,right,top,bottom)
        overwork_hours_weekday = working_hours_of_workers_sum_of_all_data[id][0][1]
        ws.cell(column = 10,row = count, value=overwork_hours_weekday)
        ws.cell(count,10).border = Border(left,right,top,bottom)
        number_holiday_day = working_hours_of_workers_sum_of_all_data[id][1][0]
        ws.cell(column = 11,row = count, value=number_holiday_day)
        ws.cell(count,11).border = Border(left,right,top,bottom)
        ws.cell(column = 11,row = count).alignment = Alignment(horizontal='center')
        overwork_hours_weekend = number_holiday_day = working_hours_of_workers_sum_of_all_data[id][1][1]
        ws.cell(column = 12,row = count, value=overwork_hours_weekend)
        ws.cell(count,12).border = Border(left,right,top,bottom)
        ws.cell(column = 13,row = count, value=working_hours_of_workers_sum_of_all_data[id][2])
        ws.cell(count,13).border = Border(left,right,top,bottom)
        ws.cell(column = 13,row = count).alignment = Alignment(horizontal='center')
        ws.cell(column = 14,row = count, value=total_salary_id[id][0])
        ws.cell(count,14).border = Border(left,right,top,bottom)
        ws.cell(column = 15,row = count, value=total_salary_id[id][1])
        ws.cell(count,15).border = Border(left,right,top,bottom)
        ws.cell(column = 16,row = count, value=total_salary_id[id][2])
        ws.cell(count,16).border = Border(left,right,top,bottom)
        count += 1
    ws.auto_filter.ref = 'A1:G24' # Создает фильтр внутри excel таблицы
    list_date =list(time_table.keys())
    number_month = int(list_date[0][5:7])
    str_year = list_date[0][:4]
    name_file_excel = month_name_for_print(number_month)+ '_' + str_year + '.xlsx'
    wb.save(name_file_excel)
    print(f"Файл {name_file_excel} c общей таблицей сформирован")

